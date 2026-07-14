import os
import sqlite3
import uuid
import time
import urllib.parse
import secrets
import requests
import base64
import json
from datetime import datetime, timezone
from flask import Flask, render_template, request, jsonify, Response, session, redirect, url_for, send_file, send_from_directory
from functools import wraps
from dotenv import load_dotenv
from werkzeug.utils import secure_filename

load_dotenv()

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
parent_dir = os.path.dirname(BASE_DIR)
sibling_frontend = os.path.join(parent_dir, 'frontend')
sibling_storage = os.path.join(parent_dir, 'storage')
sibling_backstorage = os.path.join(parent_dir, 'backstorage')

if os.path.exists(sibling_frontend):
    FRONTEND_DIR = sibling_frontend
    if os.path.exists(sibling_storage):
        BACKSTORAGE_DIR = sibling_storage
    elif os.path.exists(sibling_backstorage):
        BACKSTORAGE_DIR = sibling_backstorage
    else:
        BACKSTORAGE_DIR = sibling_storage
    static_folder_path = os.path.join(FRONTEND_DIR, 'static')
    if not os.path.exists(static_folder_path):
        static_folder_path = FRONTEND_DIR
    app = Flask(
        __name__,
        template_folder=FRONTEND_DIR,
        static_folder=static_folder_path,
        static_url_path='/static'
    )
else:
    FRONTEND_DIR = BASE_DIR
    BACKSTORAGE_DIR = BASE_DIR
    template_subfolder = os.path.join(BASE_DIR, 'templates')
    if os.path.exists(template_subfolder):
        app = Flask(__name__)
    else:
        app = Flask(__name__, template_folder=BASE_DIR, static_folder=BASE_DIR, static_url_path='/static')

app.secret_key = os.environ.get("FLASK_SECRET_KEY", secrets.token_hex(24))
DB_FILE = os.path.join(BACKSTORAGE_DIR, 'chat.db')

UPLOAD_FOLDER = os.path.join(BACKSTORAGE_DIR, 'uploads')
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER

SYSTEM_PROMPT = (
    "You are a professional AI Assistant.\n"
    "Always answer politely.\n"
    "Provide concise but accurate answers.\n"
    "When writing code:\n"
    "- Explain first\n"
    "- Then provide clean code\n"
    "- Use Markdown\n"
    "- Add comments\n"
    "- Follow best practices\n"
    "If uncertain, admit uncertainty instead of making up information."
)

# Connect to Groq API if key is available
groq_client = None
try:
    groq_api_key = os.environ.get("GROQ_API_KEY")
    if groq_api_key and not groq_api_key.startswith("your_"):
        from groq import Groq
        groq_client = Groq(api_key=groq_api_key)
        print("Groq Client initialized successfully.")
    else:
        print("WARNING: GROQ_API_KEY is not defined. Server will run in simulation mode.")
except Exception as e:
    print(f"Error initializing Groq client: {e}. Running in simulation mode.")


# Database helpers
def get_db():
    conn = sqlite3.connect(DB_FILE)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA foreign_keys = ON;")
    return conn


def init_db():
    with get_db() as conn:
        conn.execute("""
            CREATE TABLE IF NOT EXISTS conversations (
                id TEXT PRIMARY KEY,
                title TEXT NOT NULL,
                user_id TEXT NOT NULL,
                created_at TEXT NOT NULL
            )
        """)
        conn.execute("""
            CREATE TABLE IF NOT EXISTS messages (
                id TEXT PRIMARY KEY,
                conversation_id TEXT NOT NULL,
                role TEXT NOT NULL,
                content TEXT NOT NULL,
                rating TEXT,
                response_time REAL,
                attachments TEXT DEFAULT '[]',
                created_at TEXT NOT NULL,
                FOREIGN KEY (conversation_id) REFERENCES conversations(id) ON DELETE CASCADE
            )
        """)
        
        # Check if user_id column exists in conversations table
        cursor = conn.execute("PRAGMA table_info(conversations)")
        columns = [row[1] for row in cursor.fetchall()]
        if 'user_id' not in columns:
            conn.execute("ALTER TABLE conversations ADD COLUMN user_id TEXT DEFAULT 'guest';")
            
        # Check if share_id column exists in conversations table
        if 'share_id' not in columns:
            conn.execute("ALTER TABLE conversations ADD COLUMN share_id TEXT;")
            conn.execute("CREATE UNIQUE INDEX IF NOT EXISTS idx_conversations_share_id ON conversations(share_id);")
            
        # Check if attachments column exists in messages table
        cursor = conn.execute("PRAGMA table_info(messages)")
        columns = [row[1] for row in cursor.fetchall()]
        if 'attachments' not in columns:
            conn.execute("ALTER TABLE messages ADD COLUMN attachments TEXT DEFAULT '[]';")
            
        conn.commit()


init_db()


# Route protection & ownership helpers
def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            if request.path.startswith('/api/'):
                return jsonify({"error": "Unauthorized. Please login."}), 401
            return redirect(url_for('login_page'))
        return f(*args, **kwargs)
    return decorated_function


def verify_ownership(conn, conv_id, user_id):
    conv = conn.execute("SELECT user_id FROM conversations WHERE id = ?", (conv_id,)).fetchone()
    return conv is not None and conv["user_id"] == user_id


@app.after_request
def after_request(response):
    response.headers.add('Access-Control-Allow-Origin', '*')
    response.headers.add('Access-Control-Allow-Headers', 'Content-Type,Authorization')
    response.headers.add('Access-Control-Allow-Methods', 'GET,PUT,POST,DELETE,OPTIONS')
    return response


@app.route('/api/<path:path>', methods=['OPTIONS'])
def options_handler(path):
    return '', 204


# Web workspace routes
@app.route('/')
@login_required
def index():
    return render_template('index.html')


@app.route('/login')
def login_page():
    if 'user_id' in session:
        return redirect(url_for('index'))
        
    google_configured = bool(
        os.environ.get("GOOGLE_CLIENT_ID") and 
        os.environ.get("GOOGLE_CLIENT_SECRET") and
        not os.environ.get("GOOGLE_CLIENT_ID").startswith("your_")
    )
    return render_template('login.html', google_configured=google_configured)


@app.route('/login/guest')
def login_guest():
    session['user_id'] = f"guest_{str(uuid.uuid4())[:8]}"
    session['user_name'] = "Guest User"
    session['user_picture'] = None
    return redirect(url_for('index'))


@app.route('/login/google')
def login_google():
    client_id = os.environ.get("GOOGLE_CLIENT_ID")
    if not client_id:
        return "Google client credentials are not configured", 400
        
    oauth_state = secrets.token_hex(16)
    session['oauth_state'] = oauth_state
    
    redirect_host = request.host_url
    if not ('localhost' in request.host or '127.0.0.1' in request.host):
        redirect_host = redirect_host.replace('http://', 'https://')
        
    redirect_uri = redirect_host.rstrip('/') + '/login/google/callback'
    
    google_auth_endpoint = "https://accounts.google.com/o/oauth2/v2/auth"
    params = {
        "client_id": client_id,
        "redirect_uri": redirect_uri,
        "scope": "openid email profile",
        "response_type": "code",
        "state": oauth_state
    }
    
    auth_url = f"{google_auth_endpoint}?{urllib.parse.urlencode(params)}"
    return redirect(auth_url)


@app.route('/login/google/callback')
def login_google_callback():
    req_state = request.args.get('state')
    if not req_state or req_state != session.get('oauth_state'):
        return "OAuth security state mismatch", 400
        
    code = request.args.get('code')
    if not code:
        return "Authorization code missing", 400
        
    client_id = os.environ.get("GOOGLE_CLIENT_ID")
    client_secret = os.environ.get("GOOGLE_CLIENT_SECRET")
    
    redirect_host = request.host_url
    if not ('localhost' in request.host or '127.0.0.1' in request.host):
        redirect_host = redirect_host.replace('http://', 'https://')
    redirect_uri = redirect_host.rstrip('/') + '/login/google/callback'
    
    # Token exchange request
    token_endpoint = "https://oauth2.googleapis.com/token"
    token_data = {
        "code": code,
        "client_id": client_id,
        "client_secret": client_secret,
        "redirect_uri": redirect_uri,
        "grant_type": "authorization_code"
    }
    
    try:
        token_res = requests.post(token_endpoint, data=token_data)
        token_json = token_res.json()
        access_token = token_json.get("access_token")
        
        if not access_token:
            return f"Failed to retrieve Google token: {token_json.get('error_description', 'unknown error')}", 400
            
        # Retrieve user profile info
        userinfo_endpoint = "https://www.googleapis.com/oauth2/v3/userinfo"
        userinfo_res = requests.get(userinfo_endpoint, headers={"Authorization": f"Bearer {access_token}"})
        userinfo = userinfo_res.json()
        
        email = userinfo.get("email")
        name = userinfo.get("name", "Google User")
        picture = userinfo.get("picture")
        
        if not email:
            return "Google account profile info retrieval failed", 400
            
        session['user_id'] = email
        session['user_name'] = name
        session['user_picture'] = picture
        
        session.pop('oauth_state', None)
        return redirect(url_for('index'))
        
    except Exception as e:
        return f"Google Login callback error: {e}", 500


# Serving uploaded files
@app.route('/uploads/<filename>')
def uploaded_file(filename):
    return send_from_directory(app.config['UPLOAD_FOLDER'], filename)


# File upload API
ALLOWED_EXTENSIONS = {
    'png', 'jpg', 'jpeg', 'webp',
    'mp4', 'mov', 'avi',
    'pdf', 'docx', 'txt', 'csv', 'xlsx', 'json', 'md',
    'py', 'java', 'cpp', 'c', 'html', 'css', 'js'
}

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

@app.route('/api/upload', methods=['POST'])
@login_required
def upload_file():
    if 'file' not in request.files:
        return jsonify({"error": "No file part in request"}), 400
    file = request.files['file']
    if file.filename == '':
        return jsonify({"error": "No selected file"}), 400
    
    if file and allowed_file(file.filename):
        filename = secure_filename(file.filename)
        ext = filename.rsplit('.', 1)[1].lower() if '.' in filename else ''
        base = filename.rsplit('.', 1)[0] if '.' in filename else filename
        unique_name = f"{base}_{uuid.uuid4().hex[:8]}.{ext}" if ext else f"{base}_{uuid.uuid4().hex[:8]}"
        
        file_path = os.path.join(app.config['UPLOAD_FOLDER'], unique_name)
        file.save(file_path)
        
        file_size = os.path.getsize(file_path)
        mime_type = request.form.get('type') or file.mimetype or 'application/octet-stream'
        
        return jsonify({
            "id": str(uuid.uuid4()),
            "name": file.filename,
            "unique_name": unique_name,
            "url": f"/uploads/{unique_name}",
            "type": mime_type,
            "size": file_size
        }), 201
    return jsonify({"error": "File extension not allowed"}), 400


# Text and Document parsing utilities
import zipfile
import xml.etree.ElementTree as ET

def parse_docx(file_path):
    try:
        with zipfile.ZipFile(file_path) as docx:
            xml_content = docx.read('word/document.xml')
            root = ET.fromstring(xml_content)
            paragraphs = []
            for elem in root.iter():
                if elem.tag.endswith('t'):
                    paragraphs.append(elem.text)
            return "".join(paragraphs)
    except Exception as e:
        return f"[Error parsing DOCX: {e}]"

def parse_pdf(file_path):
    try:
        import pypdf
        reader = pypdf.PdfReader(file_path)
        text = ""
        for page in reader.pages:
            text += page.extract_text() or ""
        return text
    except Exception as e:
        return f"[Error parsing PDF: {e}]"

def parse_xlsx(file_path):
    try:
        import openpyxl
        wb = openpyxl.load_workbook(file_path, data_only=True)
        sheets_data = []
        for name in wb.sheetnames:
            sheet = wb[name]
            sheets_data.append(f"Sheet: {name}")
            for row in sheet.iter_rows(values_only=True):
                if any(row):
                    sheets_data.append(", ".join(str(cell) if cell is not None else "" for cell in row))
        return "\n".join(sheets_data)
    except Exception as e:
        return f"[Error parsing XLSX: {e}]"

def parse_file(file_path, filename):
    ext = filename.rsplit('.', 1)[1].lower() if '.' in filename else ''
    if ext in ['txt', 'csv', 'json', 'md', 'py', 'java', 'cpp', 'c', 'html', 'css', 'js']:
        try:
            with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
                return f.read()
        except Exception as e:
            return f"[Error reading file: {e}]"
    elif ext == 'pdf':
        return parse_pdf(file_path)
    elif ext == 'docx':
        return parse_docx(file_path)
    elif ext == 'xlsx':
        return parse_xlsx(file_path)
    return f"[Binary file: {filename}]"


@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login_page'))


# REST API endpoints
@app.route('/api/status', methods=['GET'])
def get_api_status():
    return jsonify({
        "groq": "online" if groq_client else "simulation"
    })


@app.route('/api/stats', methods=['GET'])
@login_required
def get_stats():
    user_id = session['user_id']
    try:
        with get_db() as conn:
            total_convs = conn.execute("SELECT COUNT(*) FROM conversations WHERE user_id = ?", (user_id,)).fetchone()[0]
            total_msgs = conn.execute(
                "SELECT COUNT(*) FROM messages WHERE conversation_id IN (SELECT id FROM conversations WHERE user_id = ?)", 
                (user_id,)
            ).fetchone()[0]
            
            today_str = datetime.now(timezone.utc).strftime('%Y-%m-%d')
            today_msgs = conn.execute(
                "SELECT COUNT(*) FROM messages WHERE substr(created_at, 1, 10) = ? AND conversation_id IN (SELECT id FROM conversations WHERE user_id = ?)", 
                (today_str, user_id)
            ).fetchone()[0]
            
        return jsonify({
            "total_conversations": total_convs,
            "total_messages": total_msgs,
            "today_messages": today_msgs
        })
    except Exception as e:
        print(f"Error fetching stats: {e}")
        return jsonify({"error": "Failed to load statistics"}), 500


@app.route('/api/conversations', methods=['GET'])
@login_required
def get_conversations():
    user_id = session['user_id']
    try:
        search_query = request.args.get('q', '').strip().lower()
        with get_db() as conn:
            if search_query:
                query = """
                    SELECT DISTINCT c.* FROM conversations c
                    LEFT JOIN messages m ON c.id = m.conversation_id
                    WHERE c.user_id = ? AND (c.title LIKE ? OR m.content LIKE ?)
                    ORDER BY c.created_at DESC
                """
                like_param = f"%{search_query}%"
                rows = conn.execute(query, (user_id, like_param, like_param)).fetchall()
            else:
                rows = conn.execute("SELECT * FROM conversations WHERE user_id = ? ORDER BY created_at DESC", (user_id,)).fetchall()
                
        conversations = [dict(row) for row in rows]
        return jsonify(conversations)
    except Exception as e:
        print(f"Error fetching conversations: {e}")
        return jsonify({"error": "Failed to retrieve conversations"}), 500


@app.route('/api/conversations', methods=['POST'])
@login_required
def create_conversation():
    user_id = session['user_id']
    try:
        body = request.get_json() or {}
        title = body.get('title', 'New Chat').strip()
        conv_id = str(uuid.uuid4())
        created_at = datetime.now(timezone.utc).isoformat().replace('+00:00', 'Z')
        
        with get_db() as conn:
            conn.execute(
                "INSERT INTO conversations (id, title, user_id, created_at) VALUES (?, ?, ?, ?)",
                (conv_id, title, user_id, created_at)
            )
            conn.commit()
            
        return jsonify({"id": conv_id, "title": title, "created_at": created_at}), 201
    except Exception as e:
        print(f"Error creating conversation: {e}")
        return jsonify({"error": "Failed to create conversation"}), 500


@app.route('/api/conversations', methods=['DELETE'])
@login_required
def clear_conversations():
    user_id = session['user_id']
    try:
        with get_db() as conn:
            conn.execute("DELETE FROM conversations WHERE user_id = ?", (user_id,))
            conn.commit()
        return jsonify({"success": True})
    except Exception as e:
        print(f"Error clearing conversations: {e}")
        return jsonify({"error": "Failed to clear conversations"}), 500


@app.route('/api/conversations/<conv_id>', methods=['PUT'])
@login_required
def rename_conversation(conv_id):
    user_id = session['user_id']
    try:
        body = request.get_json() or {}
        title = body.get('title', '').strip()
        if not title:
            return jsonify({"error": "Title is required"}), 400
            
        with get_db() as conn:
            if not verify_ownership(conn, conv_id, user_id):
                return jsonify({"error": "Forbidden"}), 403
                
            conn.execute(
                "UPDATE conversations SET title = ? WHERE id = ?",
                (title, conv_id)
            )
            conn.commit()
                
        return jsonify({"id": conv_id, "title": title})
    except Exception as e:
        print(f"Error renaming conversation: {e}")
        return jsonify({"error": "Failed to rename conversation"}), 500


@app.route('/api/conversations/<conv_id>', methods=['DELETE'])
@login_required
def delete_conversation(conv_id):
    user_id = session['user_id']
    try:
        with get_db() as conn:
            if not verify_ownership(conn, conv_id, user_id):
                return jsonify({"error": "Forbidden"}), 403
            conn.execute("DELETE FROM conversations WHERE id = ?", (conv_id,))
            conn.commit()
        return jsonify({"success": True})
    except Exception as e:
        print(f"Error deleting conversation: {e}")
        return jsonify({"error": "Failed to delete conversation"}), 500


@app.route('/api/conversations/<conv_id>/messages', methods=['GET'])
@login_required
def get_messages(conv_id):
    user_id = session['user_id']
    try:
        with get_db() as conn:
            if not verify_ownership(conn, conv_id, user_id):
                return jsonify({"error": "Forbidden"}), 403
            rows = conn.execute(
                "SELECT * FROM messages WHERE conversation_id = ? ORDER BY created_at ASC",
                (conv_id,)
            ).fetchall()
        messages = []
        for row in rows:
            m_dict = dict(row)
            try:
                m_dict["attachments"] = json.loads(m_dict.get("attachments") or "[]")
            except Exception:
                m_dict["attachments"] = []
            messages.append(m_dict)
        return jsonify(messages)
    except Exception as e:
        print(f"Error fetching messages: {e}")
        return jsonify({"error": "Failed to retrieve messages"}), 500


@app.route('/api/messages/<msg_id>/rating', methods=['POST'])
@login_required
def rate_message(msg_id):
    user_id = session['user_id']
    try:
        body = request.get_json() or {}
        rating = body.get('rating')
        if rating not in ['like', 'dislike', None]:
            return jsonify({"error": "Invalid rating value"}), 400
            
        with get_db() as conn:
            row = conn.execute(
                "SELECT conversation_id FROM messages WHERE id = ?", 
                (msg_id,)
            ).fetchone()
            if not row:
                return jsonify({"error": "Message not found"}), 404
                
            conv_id = row["conversation_id"]
            if not verify_ownership(conn, conv_id, user_id):
                return jsonify({"error": "Forbidden"}), 403
                
            conn.execute(
                "UPDATE messages SET rating = ? WHERE id = ?",
                (rating, msg_id)
            )
            conn.commit()
                
        return jsonify({"id": msg_id, "rating": rating})
    except Exception as e:
        print(f"Error rating message: {e}")
        return jsonify({"error": "Failed to rate message"}), 500


def generate_title(first_prompt):
    if not groq_client:
        words = first_prompt.strip().split()
        return " ".join(words[:3]) + ("..." if len(words) > 3 else "")
    try:
        response = groq_client.chat.completions.create(
            model="llama-3.1-8b-instant",
            messages=[
                {"role": "system", "content": "You are a professional chat title generator. Summarize the user's message in 2 to 4 words. Return ONLY the title text, with no quotes, no markdown, and no surrounding text."},
                {"role": "user", "content": first_prompt}
            ],
            max_tokens=15,
            temperature=0.3
        )
        title = response.choices[0].message.content.strip().replace('"', '').replace("'", "")
        return title if title else "New Chat"
    except Exception as e:
        print(f"Error generating title via API: {e}")
        words = first_prompt.strip().split()
        return " ".join(words[:3]) + ("..." if len(words) > 3 else "")


def _simulate_stream(prompt, attachments=None):
    attachment_desc = ""
    if attachments:
        attachment_desc = "### 📎 Uploaded Attachments Processed\n\n"
        for att in attachments:
            attachment_desc += f"- **{att['name']}** ({att['type']}, {att['size']} bytes) uploaded and analyzed successfully.\n"
        attachment_desc += "\n"

    simulated_text = (
        "### Simulated AI Assistant Response\n\n"
        f"{attachment_desc}"
        "It looks like your `GROQ_API_KEY` is not set in the `.env` file yet!\n\n"
        "To activate live completions using high-performance Groq models (like Llama-3.3), please follow these steps:\n"
        "1. Create a `.env` file based on `.env.example` in the project root.\n"
        "2. Add your Groq API Key as: `GROQ_API_KEY=your_actual_key`.\n"
        "3. Restart the Python server.\n\n"
        f"**Your prompt was**: \"{prompt}\"\n\n"
        "Here is a sample Markdown table to show what I can do:\n\n"
        "| Feature | Status | Technology |\n"
        "| :--- | :--- | :--- |\n"
        "| UI Layout | Verified | CSS Grid / Flexbox |\n"
        "| Database | SQLite | SQL backend |\n"
        "| API Stream | SSE | Flask Generators |\n\n"
        "And here is a sample code block in Python:\n"
        "```python\n"
        "def hello_world():\n"
        "    # Print a greeting\n"
        "    print(\"Hello from AI Assistant!\")\n"
        "```"
    )
    words = simulated_text.split(" ")
    for word in words:
        yield word + " "
        time.sleep(0.015)


@app.route('/api/conversations/<conv_id>/chat', methods=['POST'])
@login_required
def chat(conv_id):
    user_id = session['user_id']
    body = request.get_json() or {}
    message_content = body.get('message', '').strip()
    selected_model = body.get('model', 'llama-3.3-70b-versatile').strip()
    attachments = body.get('attachments', [])
    
    if not message_content:
        return jsonify({"error": "Message is required"}), 400
        
    try:
        user_msg_id = str(uuid.uuid4())
        created_at_user = datetime.now(timezone.utc).isoformat().replace('+00:00', 'Z')
        attachments_json = json.dumps(attachments)
        
        with get_db() as conn:
            if not verify_ownership(conn, conv_id, user_id):
                return jsonify({"error": "Forbidden"}), 403
                
            conn.execute(
                "INSERT INTO messages (id, conversation_id, role, content, attachments, created_at) VALUES (?, ?, ?, ?, ?, ?)",
                (user_msg_id, conv_id, 'user', message_content, attachments_json, created_at_user)
            )
            conn.commit()
            
            # Auto-generate title on first message
            msg_count = conn.execute("SELECT COUNT(*) FROM messages WHERE conversation_id = ?", (conv_id,)).fetchone()[0]
            if msg_count == 1:
                new_title = generate_title(message_content)
                conn.execute("UPDATE conversations SET title = ? WHERE id = ?", (new_title, conv_id))
                conn.commit()
                
            rows = conn.execute(
                "SELECT role, content FROM messages WHERE conversation_id = ? ORDER BY created_at ASC",
                (conv_id,)
            ).fetchall()
            
        history = [dict(row) for row in rows]
        
        # Multimodal and file parsing orchestration
        injected_context = ""
        image_payloads = []
        has_images = False
        
        for att in attachments:
            unique_name = att.get('unique_name')
            name = att.get('name', 'file')
            file_type = att.get('type', '')
            
            if not unique_name:
                continue
                
            file_path = os.path.join(app.config['UPLOAD_FOLDER'], unique_name)
            if not os.path.exists(file_path):
                continue
                
            if file_type.startswith('image/') or name.rsplit('.', 1)[1].lower() in ['png', 'jpg', 'jpeg', 'webp']:
                has_images = True
                try:
                    with open(file_path, "rb") as image_file:
                        encoded_string = base64.b64encode(image_file.read()).decode('utf-8')
                    image_payloads.append({
                        "type": "image_url",
                        "image_url": {
                            "url": f"data:{file_type};base64,{encoded_string}"
                        }
                    })
                except Exception as e:
                    print(f"Error encoding image: {e}")
            elif file_type.startswith('video/') or name.rsplit('.', 1)[1].lower() in ['mp4', 'mov', 'avi']:
                injected_context += f"\n\n[Attachment Video File: {name}]\nSimulated description: Video shows tutorial recording or session walkthrough. Format: {file_type}. Size: {os.path.getsize(file_path)} bytes.\n"
            else:
                file_text = parse_file(file_path, name)
                injected_context += f"\n\n--- Content of attachment: {name} ---\n{file_text}\n-----------------------------------\n"
                
        groq_messages = [{"role": "system", "content": SYSTEM_PROMPT}]
        for msg in history[:-1]:
            groq_messages.append({"role": msg["role"], "content": msg["content"]})
            
        final_prompt = message_content
        if injected_context:
            final_prompt = f"{final_prompt}\n{injected_context}"
            
        current_model = selected_model
        if has_images:
            current_model = "llama-3.2-11b-vision-preview"
            
        if has_images and groq_client:
            content_blocks = [{"type": "text", "text": final_prompt}]
            for img_payload in image_payloads:
                content_blocks.append(img_payload)
            groq_messages.append({"role": "user", "content": content_blocks})
        else:
            groq_messages.append({"role": "user", "content": final_prompt})
            
        def sse_generator():
            start_time = time.time()
            ai_msg_id = str(uuid.uuid4())
            response_chunks = []
            
            try:
                if groq_client:
                    stream = groq_client.chat.completions.create(
                        model=current_model,
                        messages=groq_messages,
                        stream=True,
                        temperature=0.7
                    )
                    for chunk in stream:
                        content_chunk = chunk.choices[0].delta.content or ""
                        if content_chunk:
                            response_chunks.append(content_chunk)
                            yield f"data: {content_chunk}\n\n"
                else:
                    for chunk in _simulate_stream(message_content, attachments):
                        response_chunks.append(chunk)
                        yield f"data: {chunk}\n\n"
                
                end_time = time.time()
                elapsed = round(end_time - start_time, 2)
                ai_response_content = "".join(response_chunks)
                created_at_ai = datetime.now(timezone.utc).isoformat().replace('+00:00', 'Z')
                
                with get_db() as conn:
                    conn.execute(
                        "INSERT INTO messages (id, conversation_id, role, content, response_time, created_at) VALUES (?, ?, ?, ?, ?, ?)",
                        (ai_msg_id, conv_id, 'assistant', ai_response_content, elapsed, created_at_ai)
                    )
                    conn.commit()
                    
                yield f"data: [STATS]{elapsed}\n\n"
                
            except GeneratorExit:
                end_time = time.time()
                elapsed = round(end_time - start_time, 2)
                ai_response_content = "".join(response_chunks)
                created_at_ai = datetime.now(timezone.utc).isoformat().replace('+00:00', 'Z')
                
                if ai_response_content.strip():
                    with get_db() as conn:
                        conn.execute(
                            "INSERT INTO messages (id, conversation_id, role, content, response_time, created_at) VALUES (?, ?, ?, ?, ?, ?)",
                            (ai_msg_id, conv_id, 'assistant', ai_response_content, elapsed, created_at_ai)
                        )
                        conn.commit()
                print(f"Client disconnected early for session {conv_id}.")
                
            except Exception as e:
                print(f"Error in streaming: {e}")
                yield f"data: [ERROR]Streaming failed: {str(e)}\n\n"
            finally:
                yield "data: [DONE]\n\n"
                
        return Response(sse_generator(), mimetype='text/event-stream')
        
    except Exception as e:
        print(f"Error in chat endpoint: {e}")
        return jsonify({"error": "Failed to start chat session"}), 500


@app.route('/api/conversations/<conv_id>/regenerate', methods=['POST'])
@login_required
def regenerate_response(conv_id):
    user_id = session['user_id']
    body = request.get_json() or {}
    selected_model = body.get('model', 'llama-3.3-70b-versatile').strip()
    
    try:
        with get_db() as conn:
            if not verify_ownership(conn, conv_id, user_id):
                return jsonify({"error": "Forbidden"}), 403
                
            rows = conn.execute(
                "SELECT * FROM messages WHERE conversation_id = ? ORDER BY created_at ASC",
                (conv_id,)
            ).fetchall()
            
        messages = [dict(row) for row in rows]
        if not messages:
            return jsonify({"error": "No messages to regenerate"}), 400
            
        last_user_index = -1
        for i in range(len(messages) - 1, -1, -1):
            if messages[i]["role"] == 'user':
                last_user_index = i
                break
                
        if last_user_index == -1:
            return jsonify({"error": "No user message found to respond to"}), 400
            
        last_user_msg = messages[last_user_index]["content"]
        last_user_attachments_str = messages[last_user_index].get("attachments") or "[]"
        try:
            attachments = json.loads(last_user_attachments_str)
        except Exception:
            attachments = []
            
        with get_db() as conn:
            conn.execute(
                "DELETE FROM messages WHERE conversation_id = ? AND created_at > ?",
                (conv_id, messages[last_user_index]["created_at"])
            )
            conn.commit()
            
            rows = conn.execute(
                "SELECT role, content FROM messages WHERE conversation_id = ? ORDER BY created_at ASC",
                (conv_id,)
            ).fetchall()
            
        history = [dict(row) for row in rows]
        
        # Multimodal and file parsing orchestration
        injected_context = ""
        image_payloads = []
        has_images = False
        
        for att in attachments:
            unique_name = att.get('unique_name')
            name = att.get('name', 'file')
            file_type = att.get('type', '')
            
            if not unique_name:
                continue
                
            file_path = os.path.join(app.config['UPLOAD_FOLDER'], unique_name)
            if not os.path.exists(file_path):
                continue
                
            if file_type.startswith('image/') or name.rsplit('.', 1)[1].lower() in ['png', 'jpg', 'jpeg', 'webp']:
                has_images = True
                try:
                    with open(file_path, "rb") as image_file:
                        encoded_string = base64.b64encode(image_file.read()).decode('utf-8')
                    image_payloads.append({
                        "type": "image_url",
                        "image_url": {
                            "url": f"data:{file_type};base64,{encoded_string}"
                        }
                    })
                except Exception as e:
                    print(f"Error encoding image: {e}")
            elif file_type.startswith('video/') or name.rsplit('.', 1)[1].lower() in ['mp4', 'mov', 'avi']:
                injected_context += f"\n\n[Attachment Video File: {name}]\nSimulated description: Video shows tutorial recording or session walkthrough. Format: {file_type}. Size: {os.path.getsize(file_path)} bytes.\n"
            else:
                file_text = parse_file(file_path, name)
                injected_context += f"\n\n--- Content of attachment: {name} ---\n{file_text}\n-----------------------------------\n"
                
        groq_messages = [{"role": "system", "content": SYSTEM_PROMPT}]
        for msg in history[:-1]:
            groq_messages.append({"role": msg["role"], "content": msg["content"]})
            
        final_prompt = last_user_msg
        if injected_context:
            final_prompt = f"{final_prompt}\n{injected_context}"
            
        current_model = selected_model
        if has_images:
            current_model = "llama-3.2-11b-vision-preview"
            
        if has_images and groq_client:
            content_blocks = [{"type": "text", "text": final_prompt}]
            for img_payload in image_payloads:
                content_blocks.append(img_payload)
            groq_messages.append({"role": "user", "content": content_blocks})
        else:
            groq_messages.append({"role": "user", "content": final_prompt})
            
        def sse_generator():
            start_time = time.time()
            ai_msg_id = str(uuid.uuid4())
            response_chunks = []
            
            try:
                if groq_client:
                    stream = groq_client.chat.completions.create(
                        model=current_model,
                        messages=groq_messages,
                        stream=True,
                        temperature=0.7
                    )
                    for chunk in stream:
                        content_chunk = chunk.choices[0].delta.content or ""
                        if content_chunk:
                            response_chunks.append(content_chunk)
                            yield f"data: {content_chunk}\n\n"
                else:
                    for chunk in _simulate_stream(last_user_msg, attachments):
                        response_chunks.append(chunk)
                        yield f"data: {chunk}\n\n"
                        
                end_time = time.time()
                elapsed = round(end_time - start_time, 2)
                ai_response_content = "".join(response_chunks)
                created_at_ai = datetime.now(timezone.utc).isoformat().replace('+00:00', 'Z')
                
                with get_db() as conn:
                    conn.execute(
                        "INSERT INTO messages (id, conversation_id, role, content, response_time, created_at) VALUES (?, ?, ?, ?, ?, ?)",
                        (ai_msg_id, conv_id, 'assistant', ai_response_content, elapsed, created_at_ai)
                    )
                    conn.commit()
                    
                yield f"data: [STATS]{elapsed}\n\n"
                
            except GeneratorExit:
                end_time = time.time()
                elapsed = round(end_time - start_time, 2)
                ai_response_content = "".join(response_chunks)
                created_at_ai = datetime.now(timezone.utc).isoformat().replace('+00:00', 'Z')
                
                if ai_response_content.strip():
                    with get_db() as conn:
                        conn.execute(
                            "INSERT INTO messages (id, conversation_id, role, content, response_time, created_at) VALUES (?, ?, ?, ?, ?, ?)",
                            (ai_msg_id, conv_id, 'assistant', ai_response_content, elapsed, created_at_ai)
                        )
                        conn.commit()
                print(f"Client disconnected early for session {conv_id} during regeneration.")
                
            except Exception as e:
                print(f"Error in regeneration stream: {e}")
                yield f"data: [ERROR]Regeneration failed: {str(e)}\n\n"
            finally:
                yield "data: [DONE]\n\n"
                
        return Response(sse_generator(), mimetype='text/event-stream')
        
    except Exception as e:
        print(f"Error in regenerate endpoint: {e}")
        return jsonify({"error": "Failed to start regeneration"}), 500


@app.route('/api/conversations/<conv_id>/export', methods=['GET'])
@login_required
def export_chat(conv_id):
    user_id = session['user_id']
    try:
        with get_db() as conn:
            if not verify_ownership(conn, conv_id, user_id):
                return jsonify({"error": "Forbidden"}), 403
                
            conversation = conn.execute("SELECT title FROM conversations WHERE id = ?", (conv_id,)).fetchone()
            rows = conn.execute(
                "SELECT role, content, response_time, created_at FROM messages WHERE conversation_id = ? ORDER BY created_at ASC",
                (conv_id,)
            ).fetchall()
            
        title = conversation["title"]
        messages = [dict(row) for row in rows]
        
        export_lines = [
            f"=== CONVERSATION: {title} ===",
            f"Exported at: {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M:%S UTC')}\n"
        ]
        
        for msg in messages:
            role_label = "USER" if msg["role"] == 'user' else "AI ASSISTANT"
            timestamp = msg["created_at"]
            content = msg["content"]
            
            export_lines.append(f"[{timestamp}] {role_label}:")
            export_lines.append(content)
            if msg["response_time"] is not None:
                export_lines.append(f"(Response generated in {msg['response_time']} seconds)")
            export_lines.append("-" * 40)
            
        export_text = "\n".join(export_lines)
        filename = f"{title.lower().replace(' ', '_')[:30]}_chat.txt"
        return Response(
            export_text,
            mimetype="text/plain",
            headers={"Content-disposition": f"attachment; filename={filename}"}
        )
    except Exception as e:
        print(f"Error exporting conversation: {e}")
        return jsonify({"error": "Failed to export conversation"}), 500


@app.route('/api/conversations/<conv_id>/share', methods=['POST'])
@login_required
def share_conversation(conv_id):
    user_id = session['user_id']
    try:
        with get_db() as conn:
            if not verify_ownership(conn, conv_id, user_id):
                return jsonify({"error": "Forbidden"}), 403
            
            row = conn.execute("SELECT share_id FROM conversations WHERE id = ?", (conv_id,)).fetchone()
            if not row:
                return jsonify({"error": "Conversation not found"}), 404
                
            share_id = row["share_id"]
            if not share_id:
                share_id = uuid.uuid4().hex
                conn.execute("UPDATE conversations SET share_id = ? WHERE id = ?", (share_id, conv_id))
                conn.commit()
                
        # Generate share URL using current request host
        share_url = f"{request.host_url.rstrip('/')}/share/{share_id}"
        return jsonify({"share_id": share_id, "share_url": share_url})
    except Exception as e:
        print(f"Error sharing conversation: {e}")
        return jsonify({"error": "Failed to share conversation"}), 500


@app.route('/share/<share_id>', methods=['GET'])
def view_shared_conversation(share_id):
    try:
        with get_db() as conn:
            conv = conn.execute("SELECT id, title FROM conversations WHERE share_id = ?", (share_id,)).fetchone()
            if not conv:
                return "Shared conversation not found", 404
                
            conv_id = conv["id"]
            title = conv["title"]
            
            rows = conn.execute(
                "SELECT * FROM messages WHERE conversation_id = ? ORDER BY created_at ASC",
                (conv_id,)
            ).fetchall()
            
        messages = []
        for row in rows:
            m_dict = dict(row)
            try:
                m_dict["attachments"] = json.loads(m_dict.get("attachments") or "[]")
            except Exception:
                m_dict["attachments"] = []
            messages.append(m_dict)
            
        return render_template("share_view.html", title=title, messages=messages)
    except Exception as e:
        print(f"Error retrieving shared conversation: {e}")
        return "Internal server error reading share logs", 500


@app.route('/api/conversations/<conv_id>/export_html', methods=['GET'])
@login_required
def export_chat_html(conv_id):
    from html import escape as escape_html
    user_id = session['user_id']
    try:
        with get_db() as conn:
            if not verify_ownership(conn, conv_id, user_id):
                return jsonify({"error": "Forbidden"}), 403
                
            conversation = conn.execute("SELECT title FROM conversations WHERE id = ?", (conv_id,)).fetchone()
            if not conversation:
                return jsonify({"error": "Conversation not found"}), 404
                
            rows = conn.execute(
                "SELECT * FROM messages WHERE conversation_id = ? ORDER BY created_at ASC",
                (conv_id,)
            ).fetchall()
            
        title = conversation["title"]
        messages = []
        for row in rows:
            m_dict = dict(row)
            try:
                m_dict["attachments"] = json.loads(m_dict.get("attachments") or "[]")
            except Exception:
                m_dict["attachments"] = []
            messages.append(m_dict)
            
        # We read static/style.css to inline the style!
        css_content = ""
        css_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'static', 'style.css')
        if os.path.exists(css_path):
            with open(css_path, 'r', encoding='utf-8', errors='ignore') as css_file:
                css_content = css_file.read()
                
        # Generate the fully self-contained HTML body
        messages_html = ""
        for msg in messages:
            avatar = "User" if msg["role"] == 'user' else "AI"
            attachments_list = ""
            if msg["attachments"]:
                attachments_list += '<div class="message-attachments">'
                for att in msg["attachments"]:
                    attachments_list += f"""
                    <div class="message-attachment-item">
                        <span>📎 {att['name']}</span>
                    </div>
                    """
                attachments_list += '</div>'
                
            timer_html = ""
            if msg["role"] == 'assistant' and msg["response_time"]:
                timer_html = f'<div style="font-size:0.72rem;color:#666;margin-top:8px;">Generated in {msg["response_time"]}s</div>'
                
            messages_html += f"""
            <div class="message-bubble {msg['role']}">
                <div class="message-avatar"><span>{avatar}</span></div>
                <div class="message-content-wrapper">
                    <div class="message-bubble-body markdown-body">{escape_html(msg['content'])}</div>
                    {attachments_list}
                    {timer_html}
                </div>
            </div>
            """
            
        # Self-contained template code
        html_export = f"""<!DOCTYPE html>
<html lang="en" data-theme="dark">
<head>
    <meta charset="UTF-8">
    <title>Exported Chat: {title}</title>
    <link href="https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700&display=swap" rel="stylesheet">
    <script src="https://cdn.jsdelivr.net/npm/marked/marked.min.js"></script>
    <style>
        {css_content}
        body {{ overflow-y: auto !important; height: auto !important; padding: 20px; }}
        .main-container {{ max-width: 760px; margin: 0 auto; }}
        .share-header {{ border-bottom: 1px solid #2A2A2A; padding: 24px 0; margin-bottom: 24px; }}
    </style>
</head>
<body>
    <div class="main-container">
        <header class="share-header">
            <h1 style="font-size: 1.5rem; font-weight: 700; margin-bottom: 6px; color:#fff;">{title}</h1>
            <div style="font-size: 0.78rem; color:#666;">Self-contained offline transcript exported at {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M:%S UTC')}</div>
        </header>
        <section class="messages-container" style="display:block;">
            <div class="messages-list">
                {messages_html}
            </div>
        </section>
    </div>
    <script>
        document.querySelectorAll('.markdown-body').forEach(el => {{
            const raw = el.textContent.trim();
            const isUser = el.closest('.message-bubble').classList.contains('user');
            if (isUser) {{
                el.innerHTML = escapeHtml(raw).replace(/\\n/g, '<br>');
            }} else {{
                if (typeof marked !== 'undefined') {{
                    el.innerHTML = marked.parse(raw);
                }}
            }}
        }});
        function escapeHtml(text) {{
            return text
                .replace(/&/g, "&amp;")
                .replace(/</g, "&lt;")
                .replace(/>/g, "&gt;")
                .replace(/"/g, "&quot;")
                .replace(/'/g, "&#039;");
        }}
    </script>
</body>
</html>"""
        
        filename = f"chat_export_{conv_id[:8]}.html"
        return Response(
            html_export,
            mimetype="text/html",
            headers={"Content-disposition": f"attachment; filename={filename}"}
        )
    except Exception as e:
        print(f"Error exporting conversation HTML: {e}")
        return jsonify({"error": "Failed to export conversation HTML"}), 500


@app.route('/api/conversations/export_all', methods=['GET'])
@login_required
def export_all_chats():
    user_id = session['user_id']
    user_name = session['user_name']
    try:
        with get_db() as conn:
            conversations_rows = conn.execute(
                "SELECT * FROM conversations WHERE user_id = ? ORDER BY created_at DESC", 
                (user_id,)
            ).fetchall()
            
            export_blocks = [
                f"==================================================",
                f"AI ASSISTANT - COMPLETE CHAT LOG EXPORT",
                f"User Profile: {user_name} ({user_id})",
                f"Export Date: {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M:%S UTC')}",
                f"Total Chats: {len(conversations_rows)}",
                f"==================================================\n"
            ]
            
            for conv in conversations_rows:
                conv_id = conv["id"]
                title = conv["title"]
                created_at = conv["created_at"]
                
                export_blocks.append(f"\n==================================================")
                export_blocks.append(f"CHAT TITLE: {title}")
                export_blocks.append(f"Created At: {created_at}")
                export_blocks.append(f"==================================================\n")
                
                messages_rows = conn.execute(
                    "SELECT role, content, response_time, created_at FROM messages WHERE conversation_id = ? ORDER BY created_at ASC",
                    (conv_id,)
                ).fetchall()
                
                for msg in messages_rows:
                    role_label = "USER" if msg["role"] == 'user' else "AI ASSISTANT"
                    export_blocks.append(f"[{msg['created_at']}] {role_label}:")
                    export_blocks.append(msg["content"])
                    if msg["response_time"] is not None:
                        export_blocks.append(f"(Response generated in {msg['response_time']} seconds)")
                    export_blocks.append("-" * 30)
                    
        export_text = "\n".join(export_blocks)
        filename = f"all_chats_{user_id.replace('@', '_').replace('.', '_')}.txt"
        return Response(
            export_text,
            mimetype="text/plain",
            headers={"Content-disposition": f"attachment; filename={filename}"}
        )
    except Exception as e:
        print(f"Error exporting all conversations: {e}")
        return jsonify({"error": "Failed to export all conversations"}), 500


@app.route('/download/source_code', methods=['GET'])
def download_source_code():
    try:
        base_dir = os.path.dirname(os.path.abspath(__file__))
        zip_path = os.path.join(base_dir, 'source_code.zip')
        
        if not os.path.exists(zip_path):
            zip_path = os.path.join(os.path.dirname(base_dir), 'source_code.zip')
            
        if os.path.exists(zip_path):
            return send_file(zip_path, as_attachment=True, download_name='source_code.zip')
        else:
            return "Source code zip not found. Run Compress-Archive on python backend.", 404
    except Exception as e:
        return f"Error downloading source code: {e}", 500


if __name__ == '__main__':
    port = int(os.environ.get("PORT", 5000))
    print("==================================================")
    print("AI Assistant Flask Server Starting...")
    print(f"Local Address: http://localhost:{port}")
    print("==================================================")
    app.run(host='0.0.0.0', port=port, debug=True)
